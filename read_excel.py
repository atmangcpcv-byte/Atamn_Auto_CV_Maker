import openpyxl

wb = openpyxl.load_workbook(r'c:\Auto_CV_Maker\Database\CV_DB (2).xlsx')
print('Sheets:', wb.sheetnames)

for name in wb.sheetnames:
    ws = wb[name]
    rows = list(ws.rows)
    if rows:
        print(f'\n=== Sheet: {name} ===')
        print('Cols:', [c.value for c in rows[0]])
        for row in rows[1:4]:
            print([cell.value for cell in row])
